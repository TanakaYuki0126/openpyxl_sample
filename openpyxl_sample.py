import openpyxl
#python3で動作


#Excelファイルの新規作成
#wb = openpyxl.Workbook()
#Excelファイルの読み込み
wb = openpyxl.load_workbook("/Users/yuki/Documents/test/openpyxl_sample.xlsx")


#シートの取得
#ws = wb["Sheet1"]
#先頭のシート(インデックスは0から)
ws = wb.worksheets[0]
#print(wb.index(ws))

#シート名のリスト
print(wb.sheetnames)

#シート名の確認
print(ws.title)

#シート名の変更
ws.title = "testSheet1"

#シートの追加
ws4 = wb.create_sheet(title="Sheet4", index=2)
#シートの削除
wb.remove(ws4)
#末尾のシートを削除
#wb.remove(wb.worksheets[-1])
print(wb.sheetnames)


#セルの取得
c1 = ws["A1"]
print(c1)
c2 = ws.cell(2, 3)
print(c2)

#複数のセル
rng1 = ws["A1:C3"]
print(rng1)
print(rng1[0][1])

#行指定
row1 = ws[1]
#print(row1)

#セルのアドレス指定
print(c1.coordinate)
#A1
print(c1.row)
#1
print(c1.column)
#1
print(c1.column_letter)
#A

#セルの値の読み書き(デフォルトでは数式が取得される。)
val1 = c1.value
print(val1)
#なにもないときはNone
c1.value ="aa"

#セルの書式設定
ws["B2"].number_format = "0.00"
ws["B2"].value = 120
from openpyxl.styles import Font
ws["B2"].font = Font(bold=True, italic=True)

#保存
wb.save("/Users/yuki/Documents/test/openpyxl_sample.xlsx")